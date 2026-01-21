from openpyxl.styles import Font, PatternFill, Protection, Alignment
from openpyxl.worksheet.protection import SheetProtection
from datetime import datetime
import os

def create_excel_summary(excel_path, questions, paper_sheet, serial_number, total_marks_obtained, total_possible_marks, time_taken):
    """
    Create a password-protected Excel summary file with evaluation results
    """
    try:
        # Create a new Excel workbook
        wb = Workbook()
        
        # Remove default sheet and create our sheets
        wb.remove(wb.active)
        
        # Create summary sheet
        ws_summary = wb.create_sheet("Summary", 0)
        
        # Create detailed entries sheet
        ws_entries = wb.create_sheet("Detailed Entries", 1)
        
        # Password for Excel file
        excel_password = "70186"
        
        # === SUMMARY SHEET ===
        # Header styling
        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Title
        ws_summary.merge_cells('A1:D1')
        ws_summary['A1'] = "EVALUATION SUMMARY"
        ws_summary['A1'].font = Font(bold=True, size=16, color="FFFFFF")
        ws_summary['A1'].fill = header_fill
        ws_summary['A1'].alignment = Alignment(horizontal="center")
        
        # Serial Number
        ws_summary['A3'] = "Serial Number:"
        ws_summary['B3'] = serial_number
        ws_summary['A3'].font = Font(bold=True)
        
        # Username
        ws_summary['A4'] = "Evaluator:"
        ws_summary['B4'] = os.path.basename(excel_path).split('_')[0]  # Extract username from filename
        ws_summary['A4'].font = Font(bold=True)
        
        # Date
        ws_summary['A5'] = "Evaluation Date:"
        ws_summary['B5'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws_summary['A5'].font = Font(bold=True)
        
        # Statistics
        ws_summary['A6'] = "STATISTICS"
        ws_summary['A6'].font = Font(bold=True, size=14)
        
        ws_summary['A7'] = "Total Questions:"
        ws_summary['B7'] = len(questions)
        ws_summary['A8'] = "Total Possible Marks:"
        ws_summary['B8'] = total_possible_marks
        ws_summary['A9'] = "Total Obtained Marks:"
        ws_summary['B9'] = total_marks_obtained
        ws_summary['A10'] = "Percentage:"
        ws_summary['B10'] = f"{(total_marks_obtained/total_possible_marks*100):.2f}%" if total_possible_marks > 0 else "0%"
        ws_summary['A11'] = "Time Taken:"
        ws_summary['B11'] = time_taken
        
        # Format statistics
        for row in range(7, 12):
            ws_summary[f'A{row}'].font = Font(bold=True)
            ws_summary[f'B{row}'].font = Font(bold=True)
        
        # Question Summary
        ws_summary['A13'] = "QUESTION SUMMARY"
        ws_summary['A13'].font = Font(bold=True, size=14)
        
        # Headers for question summary
        headers = ["Q.No", "Question", "Marks", "Obtained", "Status"]
        for col, header in enumerate(headers, 1):
            cell = ws_summary.cell(row=14, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Add question data
        for row, question in enumerate(questions, 15):
            ws_summary[f'A{row}'] = question.get('question_no', '')
            ws_summary[f'B{row}'] = question.get('question', '')
            ws_summary[f'C{row}'] = question.get('marks', 0)
            ws_summary[f'D{row}'] = question.get('obtained', 0)
            ws_summary[f'E{row}'] = question.get('status', 'NM')
        
        # Auto-adjust column widths
        for col in range(1, 6):
            max_length = 0
            column = chr(64 + col)  # Convert to A, B, C, etc.
            for row in range(1, len(questions) + 20):
                cell_value = ws_summary[f'{column}{row}'].value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            ws_summary.column_dimensions[column].width = min(max_length + 2, 50)
        
        # === DETAILED ENTRIES SHEET ===
        # Headers for detailed entries
        entry_headers = ["Page", "Question ID", "Question No", "Tool Used", "Marks", "Status"]
        for col, header in enumerate(entry_headers, 1):
            cell = ws_entries.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Add detailed entries from paper_sheet
        entry_row = 2
        for page_key, page_data in paper_sheet.items():
            if page_data:  # Only add if there's data for this page
                for entry in page_data:
                    ws_entries[f'A{entry_row}'] = page_key
                    ws_entries[f'B{entry_row}'] = entry.get('question_id', '')
                    ws_entries[f'C{entry_row}'] = entry.get('question_id', '')  # Will be updated with question number
                    ws_entries[f'D{entry_row}'] = entry.get('tool', '')
                    ws_entries[f'E{entry_row}'] = entry.get('marks', '')
                    ws_entries[f'F{entry_row}'] = 'Applied'
                    entry_row += 1
        
        # Auto-adjust column widths for entries sheet
        for col in range(1, 7):
            max_length = 0
            column = chr(64 + col)
            for row in range(1, entry_row):
                cell_value = ws_entries[f'{column}{row}'].value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            ws_entries.column_dimensions[column].width = min(max_length + 2, 30)
        
        # Protect both sheets with password
        for sheet in [ws_summary, ws_entries]:
            sheet.protection = SheetProtection(
                password=excel_password,
                sheet=True,
                objects=True,
                scenarios=True,
                formatCells=False,
                formatColumns=False,
                formatRows=False,
                insertColumns=False,
                insertRows=False,
                insertHyperlinks=False,
                deleteColumns=False,
                deleteRows=False,
                selectLockedCells=False,
                selectUnlockedCells=False
            )
        
        # Save the Excel file with password protection
        wb.save(excel_path)
        
        print(f"Excel summary created: {excel_path}")
        
    except Exception as e:
        print(f"Error creating Excel summary: {e}")
        # Don't raise the exception to avoid breaking the submission process
